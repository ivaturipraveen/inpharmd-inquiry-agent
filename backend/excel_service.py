"""Excel parsing + write-back for InpharmD MUE attachments.

The InpharmD platform attaches an .xlsx sheet to MUE inquiries. The sheet has
two relevant columns we care about:

  - "Medication/Vaccine Manufacturer"  → we READ this to auto-select targets
  - "Manufacturer Response"            → we WRITE this once a reply comes back

Both column header strings can vary slightly between sheets (case, punctuation,
extra whitespace) so we match by normalized substring.
"""
from __future__ import annotations

import csv
import io
import logging
import re
from dataclasses import dataclass
from typing import Any, Iterable, Optional
from urllib.parse import urlparse

import httpx
from openpyxl import load_workbook
from openpyxl.workbook import Workbook

log = logging.getLogger("inquiry.excel")

# How the writeback / extract paths tell xlsx and csv apart. xlsx is just a
# zip — the first bytes are the PK\x03\x04 ZIP local-file-header signature.
# Anything else, we treat as csv (covers utf-8, latin-1, with or without BOM).
_XLSX_MAGIC = b"PK\x03\x04"


def _detect_format(buf: bytes) -> str:
    """Return 'xlsx' or 'csv'. Content-Type headers and filename extensions
    are unreliable when files come from email forwarders / S3 mirrors, so we
    sniff the leading bytes instead."""
    return "xlsx" if buf[: len(_XLSX_MAGIC)] == _XLSX_MAGIC else "csv"


def _decode_csv(buf: bytes) -> str:
    """Decode CSV bytes tolerating UTF-8 (with optional BOM) and latin-1."""
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return buf.decode(enc)
        except UnicodeDecodeError:
            continue
    return buf.decode("utf-8", errors="replace")

# Header search patterns are ordered — the parser picks the FIRST tier that
# matches anywhere in the workbook. We prefer the Medication/Vaccine column
# over a generic "Manufacturer" so we don't grab e.g. "Fridge/Freezer
# Manufacturer" by accident in stability-excursion templates.
MEDICATION_NAME_HEADER_TIERS: tuple[tuple[set[str], tuple[str, ...]], ...] = (
    ({"medication", "name"}, ("medicationname", "vaccinename", "medicationvaccinename")),
    ({"vaccine", "name"}, ()),
    ({"drug", "name"}, ("drugname",)),
    ({"product", "name"}, ("productname",)),
)

PI_STORAGE_HEADER_TIERS: tuple[tuple[set[str], tuple[str, ...]], ...] = (
    ({"pi", "storage"}, ("pistorage", "pistoragedata", "pistoragetemp", "pistoragetemperature")),
    ({"storage", "condition"}, ("storagecondition", "storageconditions")),
    ({"storage", "temperature"}, ("storagetemperature", "storagetemp")),
    ({"storage", "requirement"}, ("storagerequirement", "storagerequirements")),
)

MANUFACTURER_HEADER_TIERS: tuple[tuple[set[str], tuple[str, ...]], ...] = (
    # Tier 1 — Medication/Vaccine Manufacturer (most specific)
    (
        {"medication", "manufacturer"},
        ("medicationmanufacturer", "vaccinemanufacturer", "medicationvaccinemanufacturer"),
    ),
    (
        {"vaccine", "manufacturer"},
        (),
    ),
    # Tier 2 — Drug / Product Manufacturer
    (
        {"drug", "manufacturer"},
        ("drugmanufacturer", "productmanufacturer"),
    ),
    # Tier 3 — bare "Manufacturer" / fuzzy fallbacks
    (
        {"manufacturer"},
        ("manufacturer", "mfr", "supplier", "vendor"),
    ),
)

# Response column — match exact phrase first to avoid grabbing
# "Manufacturer Stability Site Data" or similar by accident.
RESPONSE_HEADER_TIERS: tuple[tuple[set[str], tuple[str, ...]], ...] = (
    ({"manufacturer", "response"}, ("manufacturerresponse", "mfrresponse")),
    ({"response"}, ("reply",)),
)

_DOWNLOAD_TIMEOUT_SECONDS = 30


def _normalize(s: Any) -> str:
    """Lower-case alphanumeric squashed string for matching."""
    if s is None:
        return ""
    return re.sub(r"[^a-z0-9]+", "", str(s).lower())


def _tokenize(s: Any) -> set[str]:
    """Lower-case word set, used for fuzzy header / name matching."""
    if s is None:
        return set()
    return {t for t in re.split(r"[^a-z0-9]+", str(s).lower()) if t}


def _is_inpharmd_url(url: str) -> bool:
    """True only for InpharmD's app host (heroku/localhost). S3 buckets
    (theirs `inpharmd-asset.s3...` or ours `inpharmd-assistant.s3...`) are
    explicitly excluded — presigned S3 URLs already carry their own auth in
    the query string, and appending `access_token=` clobbers the signature."""
    try:
        host = (urlparse(url).hostname or "").lower()
    except Exception:
        return False
    if not host or host.endswith(".amazonaws.com"):
        return False
    return "inpharmd" in host or "mercer-inpharmd" in host


def download_excel(doc_url: str, access_token: Optional[str] = None) -> bytes:
    """Fetch the workbook bytes (xlsx or csv). Includes access_token only
    when the URL points at the InpharmD staging app — S3 URLs and public/CDN
    URLs are downloaded as-is. Appends the token as a query-string fragment
    rather than via httpx params= because params= can collide with an
    existing query (e.g. presigned signature) and corrupt the request URL."""
    final_url = doc_url
    if access_token and _is_inpharmd_url(doc_url) and "access_token=" not in doc_url:
        sep = "&" if "?" in doc_url else "?"
        final_url = f"{doc_url}{sep}access_token={access_token}"
    log.info(
        "excel.download url=%s token=%s",
        doc_url,
        "yes" if final_url is not doc_url else "no",
    )
    with httpx.Client(timeout=_DOWNLOAD_TIMEOUT_SECONDS, follow_redirects=True) as client:
        res = client.get(final_url)
        res.raise_for_status()
        return res.content


@dataclass
class ColumnLocation:
    sheet_name: str
    header_row: int  # 1-based
    col: int          # 1-based
    header_value: str


def _find_column_in_tier(
    workbook: Workbook,
    required_tokens: set[str],
    substrings: tuple[str, ...] = (),
    *,
    max_header_scan_rows: int = 20,
) -> Optional[ColumnLocation]:
    """Single-tier scan. Returns the first cell whose tokens match or whose
    normalized text contains a substring."""
    for ws in workbook.worksheets:
        for r_idx, row in enumerate(
            ws.iter_rows(min_row=1, max_row=max_header_scan_rows), start=1
        ):
            for c_idx, cell in enumerate(row, start=1):
                if cell.value is None:
                    continue
                toks = _tokenize(cell.value)
                norm = _normalize(cell.value)
                if required_tokens and required_tokens.issubset(toks):
                    return ColumnLocation(ws.title, r_idx, c_idx, str(cell.value))
                if substrings and any(s in norm for s in substrings):
                    return ColumnLocation(ws.title, r_idx, c_idx, str(cell.value))
    return None


def _find_column(
    workbook: Workbook,
    tiers: tuple[tuple[set[str], tuple[str, ...]], ...],
    *,
    max_header_scan_rows: int = 20,
) -> Optional[ColumnLocation]:
    """Walk the tiers in order. Returns the first match — so
    'Medication/Vaccine Manufacturer' wins over 'Fridge/Freezer Manufacturer'
    in a stability-excursion template."""
    for tokens, substrings in tiers:
        hit = _find_column_in_tier(
            workbook, tokens, substrings, max_header_scan_rows=max_header_scan_rows
        )
        if hit is not None:
            return hit
    return None


def _scan_headers(workbook: Workbook, *, max_rows: int = 20) -> list[str]:
    """Return every non-empty cell from the first `max_rows` of each sheet —
    used to build a helpful error message when no manufacturer column found."""
    found: list[str] = []
    for ws in workbook.worksheets:
        for r_idx, row in enumerate(
            ws.iter_rows(min_row=1, max_row=max_rows), start=1
        ):
            for c_idx, cell in enumerate(row, start=1):
                v = cell.value
                if v is None:
                    continue
                s = str(v).strip()
                if not s:
                    continue
                found.append(f"{ws.title}!R{r_idx}C{c_idx}={s[:60]}")
                if len(found) >= 60:
                    return found
    return found


@dataclass
class ExtractedRow:
    row_index: int       # 1-based Excel row
    raw_name: str        # value as it appeared in the cell
    medication_name: str = ""   # Medication/Vaccine Name column (if present)
    pi_storage: str = ""        # PI Storage Data column (if present)


def _extract_from_csv(csv_bytes: bytes) -> tuple[list[ExtractedRow], ColumnLocation]:
    """CSV equivalent of the xlsx extract. The first row is the header. We pick
    the best column match using the same MANUFACTURER_HEADER_TIERS, then walk
    the rest of the rows pulling non-empty values."""
    text = _decode_csv(csv_bytes)
    reader = csv.reader(io.StringIO(text))
    rows_raw = list(reader)
    if not rows_raw:
        raise ValueError("CSV is empty.")
    headers = rows_raw[0]
    # _norm and _toks live in MANUFACTURER_HEADER_TIERS-land; reuse the same
    # tokeniser openpyxl path uses so behavior matches.
    indexed = [(idx, _normalize(h), _tokenize(h)) for idx, h in enumerate(headers)]
    col_idx: Optional[int] = None
    for required_tokens, exact_norms in MANUFACTURER_HEADER_TIERS:
        # tier-1 exact-norm match first
        for idx, n, _toks in indexed:
            if n in exact_norms:
                col_idx = idx
                break
        if col_idx is not None:
            break
        # then token-superset match
        for idx, _n, toks in indexed:
            if required_tokens.issubset(toks):
                col_idx = idx
                break
        if col_idx is not None:
            break
    if col_idx is None:
        preview = "; ".join(h.strip() for h in headers if h.strip())[:300]
        raise ValueError(
            f"Could not find a Manufacturer column in the CSV. Headers I saw: {preview or '(none)'}"
        )

    # Find optional extra columns
    def _find_csv_col(tiers):
        for required_tokens, exact_norms in tiers:
            for idx, n, _toks in indexed:
                if n in exact_norms:
                    return idx
            for idx, _n, toks in indexed:
                if required_tokens.issubset(toks):
                    return idx
        return None

    med_col_idx = _find_csv_col(MEDICATION_NAME_HEADER_TIERS)
    pi_col_idx = _find_csv_col(PI_STORAGE_HEADER_TIERS)

    out: list[ExtractedRow] = []
    for r_idx, row in enumerate(rows_raw[1:], start=2):  # 1-based, header is row 1
        if col_idx >= len(row):
            continue
        val = (row[col_idx] or "").strip()
        if val:
            med_val = (row[med_col_idx] if med_col_idx is not None and med_col_idx < len(row) else "") or ""
            pi_val = (row[pi_col_idx] if pi_col_idx is not None and pi_col_idx < len(row) else "") or ""
            out.append(ExtractedRow(
                row_index=r_idx,
                raw_name=val,
                medication_name=med_val.strip(),
                pi_storage=pi_val.strip(),
            ))
    loc = ColumnLocation(
        sheet_name="csv",
        header_row=1,
        col=col_idx + 1,
        header_value=headers[col_idx],
    )
    return out, loc


def extract_manufacturer_rows(xlsx_bytes: bytes) -> tuple[list[ExtractedRow], ColumnLocation]:
    """Open the workbook in read-only mode and return every non-empty cell
    under the 'Medication/Vaccine Manufacturer' column.

    Handles both xlsx and csv inputs (sniffed by magic bytes).
    Raises ValueError if the column can't be found.
    """
    if _detect_format(xlsx_bytes) == "csv":
        return _extract_from_csv(xlsx_bytes)
    # First-pass uses read_only for speed. If we can't find the column we open
    # again in normal mode so _scan_headers can produce a useful error.
    wb = load_workbook(filename=io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    try:
        loc = _find_column(wb, MANUFACTURER_HEADER_TIERS)
        if loc is None:
            wb.close()
            wb2 = load_workbook(filename=io.BytesIO(xlsx_bytes), data_only=True)
            seen = _scan_headers(wb2)
            wb2.close()
            preview = "; ".join(seen[:10]) if seen else "(no header cells found)"
            raise ValueError(
                "Could not find a Manufacturer column in the workbook. "
                f"Headers I saw: {preview}"
            )
        # Optionally find Medication/Vaccine Name and PI Storage columns
        med_loc = _find_column(wb, MEDICATION_NAME_HEADER_TIERS)
        pi_loc = _find_column(wb, PI_STORAGE_HEADER_TIERS)

        ws = wb[loc.sheet_name]

        # Determine column range to iterate (span all three columns at once)
        all_cols = [loc.col]
        if med_loc and med_loc.sheet_name == loc.sheet_name:
            all_cols.append(med_loc.col)
        if pi_loc and pi_loc.sheet_name == loc.sheet_name:
            all_cols.append(pi_loc.col)
        min_col = min(all_cols)
        max_col = max(all_cols)

        rows: list[ExtractedRow] = []
        for r_idx, row_cells in enumerate(
            ws.iter_rows(min_row=loc.header_row + 1, min_col=min_col, max_col=max_col),
            start=loc.header_row + 1,
        ):
            # Map column index → cell value
            cell_map = {c.column: (str(c.value).strip() if c.value is not None else "") for c in row_cells}
            name = cell_map.get(loc.col, "")
            if not name:
                continue
            med_val = cell_map.get(med_loc.col, "") if med_loc and med_loc.sheet_name == loc.sheet_name else ""
            pi_val = cell_map.get(pi_loc.col, "") if pi_loc and pi_loc.sheet_name == loc.sheet_name else ""
            rows.append(ExtractedRow(
                row_index=r_idx,
                raw_name=name,
                medication_name=med_val,
                pi_storage=pi_val,
            ))
        return rows, loc
    finally:
        wb.close()


# ──────────────────────────── Matching ────────────────────────────


@dataclass
class ManufacturerMatch:
    row_index: int
    raw_name: str
    matched_id: Optional[int]
    matched_name: Optional[str]
    confidence: str  # "exact" | "partial" | "loose" | "none"
    medication_name: str = ""
    pi_storage: str = ""


def match_manufacturers(
    rows: Iterable[ExtractedRow],
    manufacturers: list,  # list of ManufacturerContact rows
) -> list[ManufacturerMatch]:
    """Pair each Excel name with the best matching manufacturer in our DB.

    Strategy (in order, first hit wins):
      1. Exact normalized match on manufacturer name OR parent_owner
      2. Either side is a substring of the other (normalized)
      3. All Excel words are present in the manufacturer's name+parent
    """
    # Pre-index for O(N) match
    by_norm: dict[str, Any] = {}
    indexed: list[tuple[str, set[str], Any]] = []
    for m in manufacturers:
        for source in (m.manufacturer, m.parent_owner):
            if source:
                key = _normalize(source)
                if key and key not in by_norm:
                    by_norm[key] = m
        toks_name = _tokenize(m.manufacturer)
        toks_parent = _tokenize(m.parent_owner) if m.parent_owner else set()
        indexed.append((_normalize(m.manufacturer), toks_name | toks_parent, m))

    out: list[ManufacturerMatch] = []
    for row in rows:
        nkey = _normalize(row.raw_name)
        rtoks = _tokenize(row.raw_name)
        extra = {"medication_name": row.medication_name, "pi_storage": row.pi_storage}
        if not nkey:
            out.append(ManufacturerMatch(row.row_index, row.raw_name, None, None, "none", **extra))
            continue

        # 1) exact
        m = by_norm.get(nkey)
        if m:
            out.append(
                ManufacturerMatch(
                    row.row_index, row.raw_name, m.id, m.manufacturer, "exact", **extra
                )
            )
            continue

        # 2) substring either way
        sub_hit = None
        for mnorm, mtoks, candidate in indexed:
            if not mnorm:
                continue
            if mnorm in nkey or nkey in mnorm:
                sub_hit = candidate
                break
        if sub_hit:
            out.append(
                ManufacturerMatch(
                    row.row_index, row.raw_name, sub_hit.id, sub_hit.manufacturer, "partial", **extra
                )
            )
            continue

        # 3) token superset
        loose_hit = None
        for mnorm, mtoks, candidate in indexed:
            if rtoks and rtoks.issubset(mtoks):
                loose_hit = candidate
                break
        if loose_hit:
            out.append(
                ManufacturerMatch(
                    row.row_index, row.raw_name, loose_hit.id, loose_hit.manufacturer, "loose", **extra
                )
            )
            continue

        out.append(ManufacturerMatch(row.row_index, row.raw_name, None, None, "none", **extra))
    return out


# ─────────────────────────── Writeback ───────────────────────────


def _write_response_csv(
    csv_bytes: bytes, *, row_index: int, response_text: str
) -> bytes:
    """CSV equivalent of write_response. Locates or appends the 'Manufacturer
    Response' column, sets the cell at row_index (1-based; header is row 1),
    and returns the re-serialised CSV. Output is always UTF-8 with CRLF row
    endings — what Excel expects when downloading."""
    text = _decode_csv(csv_bytes)
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        raise ValueError("CSV is empty; nothing to write into.")
    headers = rows[0]
    indexed = [(idx, _normalize(h), _tokenize(h)) for idx, h in enumerate(headers)]
    col_idx: Optional[int] = None
    for required_tokens, exact_norms in RESPONSE_HEADER_TIERS:
        for idx, n, _toks in indexed:
            if n in exact_norms:
                col_idx = idx
                break
        if col_idx is not None:
            break
        for idx, _n, toks in indexed:
            if required_tokens.issubset(toks):
                col_idx = idx
                break
        if col_idx is not None:
            break
    if col_idx is None:
        col_idx = len(headers)
        headers.append("Manufacturer Response")
        rows[0] = headers
        # Pad any short data rows so column indexing stays consistent.
        for r in rows[1:]:
            while len(r) < len(headers):
                r.append("")
    target = row_index - 1  # 1-based → 0-based
    if target < 1 or target >= len(rows):
        raise ValueError(
            f"row_index {row_index} out of range for CSV with {len(rows)} rows"
        )
    while len(rows[target]) <= col_idx:
        rows[target].append("")
    rows[target][col_idx] = response_text
    buf = io.StringIO()
    writer = csv.writer(buf, lineterminator="\r\n")
    writer.writerows(rows)
    return buf.getvalue().encode("utf-8")


def write_response(
    xlsx_bytes: bytes,
    *,
    row_index: int,
    response_text: str,
    sheet_name: Optional[str] = None,
) -> bytes:
    """Open the workbook, locate (or create) the 'Manufacturer Response' column,
    write `response_text` into the cell at `row_index`, and return the updated
    bytes. Pure in-memory — caller decides where to upload.

    Handles both xlsx and csv inputs (sniffed by magic bytes). For CSV the
    sheet_name argument is ignored — CSV has no sheets."""
    if _detect_format(xlsx_bytes) == "csv":
        return _write_response_csv(
            xlsx_bytes, row_index=row_index, response_text=response_text
        )
    wb = load_workbook(filename=io.BytesIO(xlsx_bytes), data_only=False)
    try:
        loc = _find_column(wb, RESPONSE_HEADER_TIERS)
        if loc is None:
            # Fall back to the same sheet as the Manufacturer column, add a
            # new "Manufacturer Response" column to the right.
            mcol = _find_column(wb, MANUFACTURER_HEADER_TIERS)
            if mcol is None:
                raise ValueError(
                    "Cannot find a 'Manufacturer Response' column and no anchor"
                    " 'Manufacturer' column to append next to."
                )
            ws = wb[sheet_name or mcol.sheet_name]
            new_col = (ws.max_column or mcol.col) + 1
            ws.cell(row=mcol.header_row, column=new_col, value="Manufacturer Response")
            ws.cell(row=row_index, column=new_col, value=response_text)
        else:
            ws = wb[sheet_name or loc.sheet_name]
            ws.cell(row=row_index, column=loc.col, value=response_text)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
    finally:
        wb.close()
